"""
NOTICE : This is just a practice project and NOT FOR COMMERCIAL USE.
Read the rules for web scraping from the MoneyControl's website (robots.txt or whatever) and scrape data AT YOUR OWN RISK.
"""

""" 
NOTe
Once script completes, find-replace .* to & using regex option 
in the new Workbook to counter invisible apostrephe problem
 """

import os
import sys
import re
import selenium
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from bs4 import BeautifulSoup
import highlight_sel_element
import xpath_soup
from openpyxl import load_workbook

import re

# iterates over the query list to retrieve data for all years spread across various pages
def iterate_over_data_pages(list_of_query_titles,year,start_col_index) :
    is_last_query_on_page = False
    size_of_list = len(list_of_query_titles)
    data_soup = BeautifulSoup(driver.page_source,"html")
    count = 0
    if False : 
        #no data on page, return
        print("HELLO, we will return from here")
    for query in list_of_query_titles:
        print("size of list : " +  str(size_of_list) +  " index : " + str(list_of_query_titles.index(query)) )
        if(list_of_query_titles.index(query) == size_of_list-1):
            is_last_query_on_page = True
        # *******
        # query_title = data_soup.find("td",text = query["title"])
        query_title = None
        for option in query[0] :
            query_title = data_soup.find("td",text = option)
            if(query_title != None) :
                break
        result_list = retrieve_data(query_title,is_last_query_on_page)
        count=store_data(result_list,query[1],year,start_col_index)
    start_col_index = start_col_index + count

    # if 'PreviousYears>>' link not found then return cuz this should be the last page
    try :
        previous_years_button = driver.find_element_by_link_text("Previous Years »")
    except :
        print("Previous Years » link not found")
        return

    # Alert may obscure PreviousYears>> link, so first remove it before clicking the link
    try :
        element = driver.find_element_by_class_name('shAlrt')
        driver.execute_script("""
        var element = arguments[0];
        element.parentNode.removeChild(element);
        """, element)
        print("Alert removed, clicking button\n")
    except :
        print ("No element with class shAlrt found : No alert\n")

    # still keeping this check in case it breaks in future, we'll know 
    try :
        previous_years_button.click()
    except Exception as ex:
        print(ex)
        print("\n\ncould not click PreviousYears>>, hence return with perhaps incomplete data\n\n")
        return
    iterate_over_data_pages(list_of_query_titles,year,start_col_index)
    return
    # go back
    # driver.execute_script("window.history.go(-1)")

def retrieve_data(query_title,is_last_query_on_page) :
    result_list = []
    curr = query_title
    print(query_title)
    while(curr!=None and curr.next_sibling != None) : #stop at last td of this query_title row
        if(curr.next_sibling.next_sibling == None) :
            #this if-else is only for informative purpose(see prints), no other use
            if(is_last_query_on_page) :
                print("no more data here")
            else :
                print("moving to the next query")
            break
        year_n = curr.next_sibling.next_sibling
        element_to_highlight = driver.find_element_by_xpath(xpath_soup.xpath_from_soup(year_n))
        highlight_sel_element.highlight(element_to_highlight)
        curr = year_n
        result_list.append(str(year_n.get_text()))
    return result_list

def store_data(result_list,row_index,year,start_col_index):
    count = 0
    if len(result_list) == 0 :
        print("result_list is empty")
        return count
    for i in range(len(result_list)):
        ws.cell(row=row_index,column=start_col_index + i, value=float(result_list[i].replace(',','')))
        print(year-i-start_col_index+2," : ",result_list[i])        #2 is hard coded as it is the seed value of start_col_index
        count = count + 1
    print("data stored\n")
    return count
    #store the result list of each query_title

if __name__ == "__main__":
    wb = load_workbook("workbooks/test_book.xlsx")
    ws = wb["Sheet1"]
    list_of_query_titles = []
    url = str(sys.argv[1])
    # create a new Firefox session
    driver = webdriver.Firefox()
    driver.implicitly_wait(30)
    driver.get(url)
    try :
        button = driver.find_element_by_link_text('FINANCIALS')
    except :
        print("element FINANCIALS cannot be found : terminating script")
        exit(0)
    button.click()

    list_of_balance_sheet_elements = [
        (["Total Shareholders Funds","Total ShareHolders Funds"],7),
        (["Long Term Borrowings"],10),
        (["Short Term Borrowings"],11),
        (["Total Current Liabilities"],16),
        (["Total Current Assets"],20),
        (["Total Assets"],19)
    ]

    list_of_ratios_elements = [
        (["Book Value [InclRevalReserve]/Share (Rs.)","Book Value [Incl. Reval Reserve]/Share (Rs.)"],3),
        (["Diluted EPS (Rs.)","Diluted Eps (Rs.)"],4),
        (["Dividend / Share(Rs.)","Dividend/Share (Rs.)"],5),
    ]

    # iteration over diff pages
    list_of_page_links = [("Balance Sheet",list_of_balance_sheet_elements),("Ratios",list_of_ratios_elements)]
    for page_name, list_of_elements in list_of_page_links : 
        year = 2018
        start_col_index = 2
        button = driver.find_element_by_link_text(page_name)
        highlight_sel_element.highlight(button)
        button.click()   
        iterate_over_data_pages(list_of_elements,year,start_col_index)
    print("saving file")
    wb.save("workbooks/" + str(sys.argv[2])+".xlsx")